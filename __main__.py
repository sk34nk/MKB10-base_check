#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon May 10 13:25:08 2021

@author: sk34nk
"""
import sys
import openpyxl

#from .DBSCAN_scan import DBSCAN_OBJ
#def DB(dict_nml):
#    DBSCAN_obj = DBSCAN_OBJ(dict_nml)
#    DBSCAN_obj.DBSCAN_create()
def get_MKB10():
    workbook = openpyxl.load_workbook('./source_data//MKB10.xlsx')
    return  workbook['Sheet1']    

def get_code(string, MKB10_list):
    string_code = 'Нет результата'
    string = string.lower()
    MKB10_list = get_MKB10()
    for row in range(1,MKB10_list.max_row):
        current_cell = str(MKB10_list.cell(row, 2).value)
        current_cell = current_cell.lower()
        if (current_cell.find(string) != -1): #вместо test ввести искомое значение
            string_code = MKB10_list.cell(row,1).value
            break
    return string_code

def main(argv):
    filename = './rez50.xlsx'
    dict_error = {}
    dict_number = {}
    print("Start reading file")
    workbook = openpyxl.load_workbook(filename)
    input_data = workbook['Лист1']
    print("File in memory")
    MKB10 = get_MKB10()
    for row in range(1, (input_data.max_row + 1)):
        string_diagnosis = input_data.cell(row, 6).value
        if not string_diagnosis is None:
            string_diagnosis = string_diagnosis.strip()
            string_diagnosis = string_diagnosis.replace('12-', 'двенадцати')
            input_data.cell(row, 6).value = string_diagnosis
            print("Getting code ", row, " string")
            input_data.cell(row, 12).value = get_code(string_diagnosis, MKB10)
    print("Start saving file")
    workbook.save(filename)
    print("File on disk")
    
    print(dict_error)
    print(dict_number)
        
if __name__ == '__main__':
    main(sys.argv)
    
    
